#!/usr/bin/env python3
"""
Daily Data Analytics / Data Engineering job scanner.

Sources:
- Greenhouse
- Lever
- Ashby
- Workday direct CXS
- Optional Workday via Apify

Outputs:
- results/data_jobs_<timestamp>.xlsx
- results/data_jobs_<timestamp>.md
- results/seen.json

Requires:
- openpyxl==3.1.5
"""

import hashlib
import html
import json
import os
import re
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RESULTS_DIR = Path("results")
SEEN_FILE = RESULTS_DIR / "seen.json"

# Data-heavy + general tech companies known to run public Greenhouse boards.
# NOT ALL SLUGS ARE GUARANTEED CORRECT -- some companies may have moved ATS,
# renamed their slug, or never used Greenhouse at all. Run the script, check
# the "Scanner Health" tab for 404 errors, and delete/fix the broken ones.
# This pruning step is a normal part of maintaining this list, not a bug.
PRIORITY_GREENHOUSE = [
    "databricks", "fivetran", "amplitude", "mixpanel", "mongodb",
    "elastic", "hightouch", "grafanalabs", "gitlab", "confluent",
    "doordash", "instacart", "pinterest", "affirm", "brex", "plaid",
    "airtable", "asana", "discord", "dropbox", "duolingo", "figma",
    "gusto", "hashicorp", "intercom", "lyft", "peloton", "reddit",
    "robinhood", "squarespace", "twilio", "zendesk", "coinbase",
    "cloudflare", "datadog", "docker", "flexport", "honeycombio",
    "okta", "pagerduty", "postmanlabs", "samsara", "stripe",
    "wealthfront", "webflow", "zapier", "benchling", "calendly",
    "carta", "checkr", "clari", "coursera", "digitalocean",
    "fastly", "faire", "front", "gong", "grammarly", "handshake",
    "harness", "hubspot", "iterable", "klaviyo", "lattice",
    "launchdarkly", "mailchimp", "marqeta", "medallia", "miro",
    "monte-carlo-data", "mural", "netlify", "nextdoor", "opsgenie",
    "outreach", "patreon", "pendo", "procore", "qualtrics", "quora",
    "redis", "rippling", "roblox", "rubrik", "samsclub", "seatgeek",
    "servicetitan", "sigmacomputing", "smartsheet", "sofi",
    "spring-health", "strava", "sumologic", "talkdesk", "thumbtack",
    "toast", "unity", "vanta", "verkada", "vimeo", "wish",
    "workiva", "yelp", "zillow", "zscaler",
]

SMARTRECRUITERS_SLUGS = [
    # "smartrecruiters" is SmartRecruiters' own careers board -- confirmed
    # real via their public API docs, safe starting point. Add more by
    # visiting https://jobs.smartrecruiters.com/<identifier> in a browser
    # to confirm the identifier before adding it here.
    "smartrecruiters",
]

WORKABLE_SLUGS = [
    # Add Workable-hosted company slugs here. Find one by visiting a
    # company's careers page -- if it's Workable-hosted, the URL looks
    # like https://apply.workable.com/<slug>/ or the page embeds a
    # Workable widget referencing that same slug. Verify in a browser
    # before adding, same as Greenhouse/SmartRecruiters slugs.
]

WORKDAY_DIRECT = [
    {"name": "Adobe", "url": "https://adobe.wd5.myworkdayjobs.com/external_experienced/jobs"},
    {"name": "Mastercard", "url": "https://mastercard.wd1.myworkdayjobs.com/CorporateCareers/jobs"},
    {"name": "PayPal", "url": "https://paypal.wd1.myworkdayjobs.com/jobs/jobs"},
    {"name": "CVS Health", "url": "https://cvshealth.wd1.myworkdayjobs.com/CVS_Health_Careers/jobs"},
    {"name": "McKesson", "url": "https://mckesson.wd3.myworkdayjobs.com/External_Careers/jobs"},
    {"name": "Nordstrom", "url": "https://nordstrom.wd501.myworkdayjobs.com/nordstrom_careers/jobs"},
]

# Add blocked Workday career-board URLs here after testing them.
# Apify will not run unless this list has companies.
WORKDAY_APIFY = [
    # {"name": "Example", "url": "https://example.wd5.myworkdayjobs.com/External/jobs"},
]

TITLE_INCLUDE = [
    "data engineer", "data analyst", "analytics engineer",
    "business intelligence engineer", "bi engineer", "bi analyst",
    "business intelligence analyst", "data platform engineer",
    "data infrastructure engineer", "etl engineer", "elt engineer",
    "data pipeline engineer", "data warehouse engineer",
    "machine learning engineer", "mlops engineer",
    "ml infrastructure engineer", "data scientist",
    "quantitative analyst", "reporting analyst", "insights analyst",
    "data quality engineer", "analytics manager", "data architect",
    "database engineer", "database administrator",
]

TITLE_EXCLUDE = [
    "intern", "internship", "new grad", "graduate", "apprentice",
    "director", "vice president", "vp ", "head of",
    "sales", "marketing", "recruiter", "talent acquisition",
    "account executive", "customer success manager", "product manager",
    "program manager", "project manager", "frontend", "front-end",
    "mobile engineer", "ios engineer", "android engineer",
    "data entry", "data entry clerk",
]

# Note: "manager" is intentionally NOT excluded globally here because
# "analytics manager" is a common IC-adjacent title in data teams.
# If you only want individual-contributor roles, add "manager" back
# to TITLE_EXCLUDE and remove "analytics manager" from TITLE_INCLUDE.

TECH_KEYWORDS = [
    "sql", "python", "airflow", "dbt", "snowflake", "databricks",
    "spark", "kafka", "bigquery", "redshift", "tableau", "looker",
    "power bi", "etl", "elt", "data warehouse", "data lake",
    "pandas", "data modeling", "dagster", "fivetran", "dagster",
    "terraform", "aws", "gcp", "azure", "postgres", "data governance",
]

SPONSORSHIP_BLOCKED = [
    "unable to sponsor", "no sponsorship", "will not sponsor",
    "cannot sponsor", "must not require sponsorship",
    "without sponsorship now or in the future",
    "without current or future sponsorship",
    "u.s. citizenship required", "us citizenship required",
    "must be a u.s. citizen", "must be a us citizen",
    "active security clearance", "security clearance required",
    "ts/sci", "secret clearance",
]

SPONSORSHIP_REVIEW = [
    "export controlled", "export control", "export authorization",
    "u.s. person", "us person", "itar", "ear regulations",
]

SPONSORSHIP_POSITIVE = [
    "visa sponsorship is available", "sponsorship is available",
    "will sponsor", "we sponsor", "h-1b sponsorship", "h1b sponsorship",
]

NON_US = [
    "canada", "united kingdom", "london", "toronto", "vancouver",
    "india", "bangalore", "bengaluru", "mumbai", "hyderabad",
    "germany", "berlin", "france", "paris", "ireland", "dublin",
    "australia", "sydney", "singapore", "japan", "tokyo",
    "netherlands", "amsterdam", "poland", "romania", "brazil",
    "mexico", "israel", "tel aviv", "spain", "portugal",
    "switzerland", "zurich", "sweden", "stockholm", "denmark",
    "copenhagen", "finland", "helsinki", "norway", "oslo",
    "czech", "prague", "hungary", "budapest", "new zealand",
    "auckland", "south africa", "johannesburg", "argentina",
    "buenos aires", "colombia", "bogota", "chile", "santiago",
    "peru", "lima", "uae", "dubai", "saudi arabia", "riyadh",
    "pakistan", "bangladesh", "philippines", "manila", "vietnam",
    "hanoi", "indonesia", "jakarta", "malaysia", "kuala lumpur",
    "thailand", "bangkok", "taiwan", "taipei", "south korea",
    "seoul", "hong kong",
]

US_LOCATION_KEYWORDS = [
    "united states", "usa", "u.s.", "us remote", "remote us", "remote - us",
    "remote - united states", "remote, united states", "remote within the us",
    "remote within the united states", "united states remote",
    "new york", "new jersey", "california", "washington state",
    "texas", "florida", "georgia", "illinois",
    "massachusetts", "virginia", "north carolina",
    "colorado", "arizona", "oregon", "pennsylvania",
    "connecticut", "maryland", "ohio", "michigan",
    "minnesota", "tennessee", "utah", "nevada",
    "wisconsin", "south carolina", "alabama", "alaska",
    "arkansas", "delaware", "hawaii", "idaho",
    "indiana", "iowa", "kansas", "kentucky",
    "louisiana", "maine", "mississippi", "missouri",
    "montana", "nebraska", "new hampshire", "new mexico",
    "north dakota", "oklahoma", "rhode island",
    "south dakota", "vermont", "west virginia", "wyoming",
    "seattle", "bellevue", "san francisco", "sunnyvale", "san jose",
    "los angeles", "long beach", "irvine", "san diego", "newark", "jersey city",
    "edison", "boston", "atlanta", "austin", "dallas", "houston", "chicago",
    "charlotte", "raleigh", "denver", "phoenix", "portland", "philadelphia",
    "stamford", "washington dc", "reston", "herndon", "mclean", "livingston",
]

# 2-letter state abbreviations (CA, WA, TX, etc.) are intentionally NOT in
# the list above, because plain substring matching on 2-letter codes causes
# false positives -- "wa" matches inside "Waterloo", "ga"/"or" match inside
# "Bangalore", "in"/"ia" match inside "INDIA". They're matched separately
# below using a word-boundary regex, which only matches a standalone
# 2-letter token (e.g. ", CA" or ", CA,"), not a substring inside a word.
US_STATE_ABBR_REGEX = re.compile(
    r"\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|HI|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI"
    r"|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT"
    r"|VA|VT|WA|WI|WV|WY)\b"
    # No IGNORECASE: state codes are always written uppercase in real
    # postings. Case-insensitive matching caused "or" (the English word,
    # e.g. "...AB, BC, or NS Only") to collide with "OR" (Oregon).
)

GLOBAL_REMOTE_BLOCKERS = [
    "worldwide", "global", "anywhere", "anywhere in the world", "emea", "apac",
    "europe", "latin america", "latam", "remote - europe", "remote europe",
    "remote - canada", "remote canada", "remote - india", "remote india",
    "remote - apac", "remote apac", "remote - emea", "remote emea",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json,text/plain,*/*",
}

HEALTH = []


def get_json(url, timeout=25):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def get_text(url, timeout=25):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read().decode("utf-8")


def post_json(url, payload, headers=None, timeout=40):
    request_headers = {**HEADERS, "Content-Type": "application/json"}
    if headers:
        request_headers.update(headers)
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers=request_headers,
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def clean_html(value):
    text = html.unescape(value or "")
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_url(url):
    if not url:
        return ""
    parsed = urllib.parse.urlsplit(url)
    clean_query = urllib.parse.urlencode([
        (k, v) for k, v in urllib.parse.parse_qsl(parsed.query)
        if not k.lower().startswith("utm_") and k.lower() not in {"source", "gh_src"}
    ])
    return urllib.parse.urlunsplit(
        (parsed.scheme, parsed.netloc, parsed.path.rstrip("/"), clean_query, "")
    )


def make_id(ats, company, native_id, url, title):
    raw = native_id or normalize_url(url) or f"{company}|{title}"
    digest = hashlib.sha256(f"{ats}|{company}|{raw}".encode()).hexdigest()[:24]
    return f"{ats.lower().replace(' ', '_')}_{digest}"


def title_matches(title, department="", description=""):
    t = (title or "").lower()
    combined = f"{title} {department} {description}".lower()

    if any(word in t for word in TITLE_EXCLUDE):
        return False

    if any(word in t for word in TITLE_INCLUDE):
        return True

    # Only fall back to keyword scoring for genuinely ambiguous titles
    # (plain "analyst") -- NOT "software engineer" or "systems engineer",
    # since those show up constantly with unrelated tech stacks and were
    # causing false positives (generic backend/infra roles slipping in).
    if t.strip() == "analyst" or "data analyst" in combined:
        return sum(word in combined for word in TECH_KEYWORDS) >= 3

    return False


def location_status(location, description=""):
    """Whitelist approach: only accept a posting as US-based if the
    LOCATION FIELD ITSELF (never the job description) contains a clear
    US indicator. Everything else defaults to non_us.

    This intentionally ignores `description` entirely. The previous
    version checked location+description together and treated any
    mention of "United States" as an override -- but job descriptions
    almost always contain generic EEO/legal boilerplate mentioning
    "United States" regardless of the actual job location, which let
    non-US postings (India, Canada, Brazil, Belgium, etc.) leak through
    the filter. Relying only on the location field, which is a
    company-curated structured field, is far more reliable.
    """
    location_text = (location or "").strip()
    loc_lower = location_text.lower()

    # A handful of well-known non-US cities whose country-code suffix
    # collides with a US state abbreviation (e.g. "Herzliya, IL" means
    # Israel, not Illinois; "Waterloo, ON" means Ontario, not Oregon).
    # Checked first so they can't be misread as US by the regex below.
    known_non_us_cities = [
        "herzliya", "tel aviv", "ra'anana", "raanana", "netanya",
        "waterloo, on", "toronto, on", "ottawa, on",
    ]
    if any(city in loc_lower for city in known_non_us_cities):
        return "non_us"

    if any(blocker in loc_lower for blocker in GLOBAL_REMOTE_BLOCKERS):
        return "non_us"

    if any(us_location in loc_lower for us_location in US_LOCATION_KEYWORDS):
        if "remote" in loc_lower:
            return "us_remote"
        return "us"

    if US_STATE_ABBR_REGEX.search(location_text):
        return "us"

    return "non_us"


def sponsorship_status(description):
    text = (description or "").lower()

    for phrase in SPONSORSHIP_POSITIVE:
        if phrase in text:
            return "eligible", phrase

    for phrase in SPONSORSHIP_BLOCKED:
        if phrase in text:
            return "blocked", phrase

    for phrase in SPONSORSHIP_REVIEW:
        if phrase in text:
            return "review_required", phrase

    return "unknown", ""


def build_job(ats, company, native_id, title, location, department, url,
              posted="", description="", employment_type=""):
    description = clean_html(description)
    status, reason = sponsorship_status(description)

    return {
        "id": make_id(ats, company, str(native_id), url, title),
        "title": (title or "").strip(),
        "company": (company or "").strip(),
        "location": (location or "").strip(),
        "department": (department or "").strip(),
        "employment_type": (employment_type or "").strip(),
        "ats": ats,
        "posted": str(posted or "").strip(),
        "url": normalize_url(url),
        "description": description,
        "location_status": location_status(location, description),
        "sponsorship_status": status,
        "sponsorship_reason": reason,
    }


def route(job):
    if not title_matches(job["title"], job["department"], job["description"]):
        return "skipped", "Title/technology mismatch"

    if job["location_status"] == "non_us":
        return "skipped", "Non-US location"

    if job["sponsorship_status"] == "blocked":
        return "blocked", f"Sponsorship restriction: {job['sponsorship_reason']}"

    if job["sponsorship_status"] == "review_required":
        return "review", f"Manual review: {job['sponsorship_reason']}"

    return "matched", ""


def add_routed(job, buckets):
    destination, reason = route(job)
    if reason:
        job["skip_reason"] = reason
    buckets[destination].append(job)


def scrape_greenhouse(slug):
    start = time.time()
    buckets = {"matched": [], "review": [], "blocked": [], "skipped": []}
    try:
        data = get_json(f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true")
        jobs = data.get("jobs", [])
        for raw in jobs:
            departments = ", ".join(
                x.get("name", "") for x in raw.get("departments", []) if x.get("name")
            )
            job = build_job(
                "Greenhouse",
                slug.replace("-", " ").title(),
                raw.get("id", ""),
                raw.get("title", ""),
                (raw.get("location") or {}).get("name", ""),
                departments,
                raw.get("absolute_url", ""),
                (raw.get("updated_at", "") or "")[:10],
                raw.get("content", ""),
            )
            add_routed(job, buckets)
        HEALTH.append(["Greenhouse", slug, "ok", len(jobs), len(buckets["matched"]),
                       len(buckets["blocked"]), len(buckets["skipped"]),
                       round(time.time() - start, 2), ""])
    except Exception as exc:
        HEALTH.append(["Greenhouse", slug, "error", 0, 0, 0, 0,
                       round(time.time() - start, 2), str(exc)])
    return buckets


def scrape_smartrecruiters(company_identifier):
    start = time.time()
    buckets = {"matched": [], "review": [], "blocked": [], "skipped": []}
    try:
        jobs = []
        offset, limit = 0, 100
        while True:
            data = get_json(
                f"https://api.smartrecruiters.com/v1/companies/{company_identifier}"
                f"/postings?limit={limit}&offset={offset}"
            )
            content = data.get("content", []) or []
            jobs.extend(content)
            total = int(data.get("totalFound", 0) or 0)
            offset += limit
            if not content or offset >= total:
                break
            time.sleep(0.2)

        for raw in jobs:
            loc = raw.get("location", {}) or {}
            location_parts = [
                loc.get("city", ""), loc.get("region", ""), loc.get("country", ""),
            ]
            location = ", ".join(p for p in location_parts if p)
            if loc.get("remote"):
                location = (location + " - Remote").strip(" -")
            department = (raw.get("department", {}) or {}).get("label", "")
            posting_id = raw.get("id", "")
            # NOTE: this list-level endpoint does not return full job
            # description text -- only /postings/{id} does, which would
            # require one extra request per job. Sponsorship detection
            # will mostly show "unknown" for this source as a result;
            # that's an accepted tradeoff, not a bug.
            job = build_job(
                "SmartRecruiters",
                company_identifier.replace("-", " ").title(),
                posting_id,
                raw.get("name", ""),
                location,
                department,
                raw.get("ref", "") or f"https://jobs.smartrecruiters.com/{company_identifier}/{posting_id}",
                (raw.get("releasedDate", "") or "")[:10],
                "",
            )
            add_routed(job, buckets)
        HEALTH.append(["SmartRecruiters", company_identifier, "ok", len(jobs),
                       len(buckets["matched"]), len(buckets["blocked"]),
                       len(buckets["skipped"]), round(time.time() - start, 2), ""])
    except Exception as exc:
        HEALTH.append(["SmartRecruiters", company_identifier, "error", 0, 0, 0, 0,
                       round(time.time() - start, 2), str(exc)])
    return buckets


def scrape_workable(slug):
    start = time.time()
    buckets = {"matched": [], "review": [], "blocked": [], "skipped": []}
    try:
        data = get_json(f"https://www.workable.com/api/accounts/{slug}?details=true")
        jobs = data.get("jobs", []) or []
        for raw in jobs:
            loc = raw.get("location", {}) or {}
            location = loc.get("location_str", "") or ", ".join(
                p for p in [loc.get("city", ""), loc.get("region", ""), loc.get("country", "")]
                if p
            )
            if raw.get("telecommuting"):
                location = (location + " - Remote").strip(" -")
            job = build_job(
                "Workable", slug.replace("-", " ").title(),
                raw.get("shortcode", "") or raw.get("id", ""),
                raw.get("title", "") or raw.get("name", ""),
                location,
                raw.get("department", "") or "",
                raw.get("url", "") or raw.get("shortlink", ""),
                (raw.get("published_on", "") or raw.get("created_at", "") or "")[:10],
                raw.get("description", "") or "",
            )
            add_routed(job, buckets)
        status = "ok" if jobs else "empty"
        HEALTH.append(["Workable", slug, status, len(jobs), len(buckets["matched"]),
                       len(buckets["blocked"]), len(buckets["skipped"]),
                       round(time.time() - start, 2),
                       "" if jobs else "Zero postings returned"])
    except Exception as exc:
        HEALTH.append(["Workable", slug, "error", 0, 0, 0, 0,
                       round(time.time() - start, 2), str(exc)])
    return buckets


def workday_parts(base_url):
    host = base_url.split("//", 1)[1].split("/", 1)[0]
    tenant = host.split(".")[0]
    wd = host.split(".")[1]
    site = base_url.split(".myworkdayjobs.com/", 1)[1].split("/", 1)[0]
    cxs = f"https://{tenant}.{wd}.myworkdayjobs.com/wday/cxs/{tenant}/{site}/jobs"
    referer = f"https://{tenant}.{wd}.myworkdayjobs.com/en-US/{site}"
    return tenant, wd, site, cxs, referer


def scrape_workday_direct(company):
    start = time.time()
    buckets = {"matched": [], "review": [], "blocked": [], "skipped": []}
    try:
        tenant, wd, site, cxs, referer = workday_parts(company["url"])
        offset, page_size = 0, 20
        jobs = []
        while True:
            data = post_json(
                cxs,
                {"appliedFacets": {}, "limit": page_size,
                 "offset": offset, "searchText": " "},
                {"Referer": referer, "Accept-Language": "en-US"},
            )
            postings = data.get("jobPostings", []) or []
            total = int(data.get("total", 0) or 0)
            if not postings:
                break
            jobs.extend(postings)
            offset += page_size
            if offset >= total:
                break
            time.sleep(0.25)

        for raw in jobs:
            path = raw.get("externalPath", "") or ""
            url = path
            if path and not path.startswith("http"):
                url = f"https://{tenant}.{wd}.myworkdayjobs.com/{site}{path}"
            job = build_job(
                "Workday", company["name"], path,
                raw.get("title", ""), raw.get("locationsText", "") or "",
                "", url, raw.get("postedOn", "") or "",
            )
            add_routed(job, buckets)

        HEALTH.append(["Workday Direct", company["name"], "ok", len(jobs),
                       len(buckets["matched"]), len(buckets["blocked"]),
                       len(buckets["skipped"]), round(time.time() - start, 2), ""])
    except Exception as exc:
        HEALTH.append(["Workday Direct", company["name"], "error", 0, 0, 0, 0,
                       round(time.time() - start, 2), str(exc)])
    return buckets


def scrape_workday_apify(company):
    start = time.time()
    buckets = {"matched": [], "review": [], "blocked": [], "skipped": []}
    token = os.environ.get("APIFY_TOKEN", "")
    if not token:
        HEALTH.append(["Workday Apify", company["name"], "skipped", 0, 0, 0, 0,
                       0, "APIFY_TOKEN missing"])
        return buckets

    actor = os.environ.get(
        "APIFY_ACTOR_ID",
        "automation-lab~workday-jobs-scraper",
    )
    try:
        run_url = (
            f"https://api.apify.com/v2/acts/{urllib.parse.quote(actor, safe='~')}/runs"
            f"?token={urllib.parse.quote(token)}&waitForFinish=120"
        )
        run = post_json(
            run_url,
            {
                "companyUrl": company["url"].replace("/jobs", ""),
                "searchQuery": "data",
                "maxJobs": 250,
                "includeDescription": True,
            },
            timeout=150,
        )
        dataset_id = run.get("data", {}).get("defaultDatasetId", "")
        if not dataset_id:
            raise RuntimeError("No Apify dataset ID returned")

        items = get_json(
            f"https://api.apify.com/v2/datasets/{dataset_id}/items"
            f"?token={urllib.parse.quote(token)}&format=json&clean=true",
            timeout=60,
        )
        jobs = items if isinstance(items, list) else []

        for raw in jobs:
            job = build_job(
                "Workday/Apify", company["name"],
                raw.get("jobId", "") or raw.get("url", ""),
                raw.get("title", "") or raw.get("positionTitle", ""),
                raw.get("location", "") or raw.get("locationsText", ""),
                raw.get("department", "") or "",
                raw.get("url", "") or raw.get("applyUrl", ""),
                raw.get("postedDate", "") or raw.get("postedOn", ""),
                raw.get("description", "") or raw.get("jobDescription", ""),
                raw.get("employmentType", "") or "",
            )
            add_routed(job, buckets)

        HEALTH.append(["Workday Apify", company["name"], "ok" if jobs else "empty",
                       len(jobs), len(buckets["matched"]), len(buckets["blocked"]),
                       len(buckets["skipped"]), round(time.time() - start, 2),
                       "" if jobs else "Zero jobs returned"])
    except Exception as exc:
        HEALTH.append(["Workday Apify", company["name"], "error", 0, 0, 0, 0,
                       round(time.time() - start, 2), str(exc)])
    return buckets


def load_seen():
    if not SEEN_FILE.exists():
        return {}
    try:
        data = json.loads(SEEN_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_seen(seen):
    RESULTS_DIR.mkdir(exist_ok=True)
    SEEN_FILE.write_text(json.dumps(dict(sorted(seen.items())), indent=2), encoding="utf-8")


def dedupe(jobs):
    return list({job["id"]: job for job in jobs}.values())


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")


def format_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for column in ws.columns:
        width = min(max(len(str(c.value or "")) for c in column) + 2, 55)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(width, 12)
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_jobs(ws, title, jobs, seen, fill=None):
    ws.title = title
    ws.append([
        "New", "Title", "Company", "Location", "Department", "ATS",
        "Posted", "Sponsorship", "Sponsorship Detail", "First Seen", "Apply",
    ])
    for job in jobs:
        ws.append([
            "YES" if job.get("is_new") else "",
            job["title"], job["company"], job["location"],
            job["department"], job["ats"], job["posted"],
            job["sponsorship_status"], job["sponsorship_reason"],
            seen.get(job["id"], ""), "Apply",
        ])
        row = ws.max_row
        link = ws.cell(row=row, column=11)
        if job["url"]:
            link.hyperlink = job["url"]
            link.style = "Hyperlink"
        if fill:
            for cell in ws[row]:
                cell.fill = fill
        elif job.get("is_new"):
            for cell in ws[row]:
                cell.fill = GREEN_FILL
    format_sheet(ws)


def write_all_roles(ws, jobs, seen):
    """One combined view of every role that passed title+location filters,
    regardless of sponsorship status -- so sponsoring and non-sponsoring
    companies are visible side by side with a status column to sort/filter by.
    """
    ws.title = "All Roles (Any Sponsorship)"
    ws.append([
        "New", "Title", "Company", "Location", "Department", "ATS",
        "Posted", "Sponsorship Status", "Sponsorship Detail",
        "First Seen", "Apply",
    ])
    sponsorship_fill = {
        "eligible": GREEN_FILL,
        "unknown": None,
        "review_required": YELLOW_FILL,
        "blocked": RED_FILL,
    }
    for job in jobs:
        ws.append([
            "YES" if job.get("is_new") else "",
            job["title"], job["company"], job["location"],
            job["department"], job["ats"], job["posted"],
            job["sponsorship_status"], job["sponsorship_reason"],
            seen.get(job["id"], ""), "Apply",
        ])
        row = ws.max_row
        link = ws.cell(row=row, column=11)
        if job["url"]:
            link.hyperlink = job["url"]
            link.style = "Hyperlink"
        fill = sponsorship_fill.get(job["sponsorship_status"])
        if fill:
            for cell in ws[row]:
                cell.fill = fill
    format_sheet(ws)


def write_skipped(ws, jobs):
    ws.title = "Skipped Roles"
    ws.append(["Title", "Company", "Location", "Department", "ATS", "Reason", "Apply"])
    for job in jobs:
        ws.append([
            job["title"], job["company"], job["location"],
            job["department"], job["ats"], job.get("skip_reason", ""), "Apply",
        ])
        link = ws.cell(row=ws.max_row, column=7)
        if job["url"]:
            link.hyperlink = job["url"]
            link.style = "Hyperlink"
    format_sheet(ws)


def write_health(ws):
    ws.title = "Scanner Health"
    ws.append([
        "Source", "Company", "Status", "Jobs Returned", "Matched",
        "Blocked", "Skipped", "Duration Seconds", "Message",
    ])
    for row in HEALTH:
        ws.append(row)
    format_sheet(ws)


def write_counts(ws, title, jobs, field):
    ws.title = title
    ws.append([field.title(), "Role Count"])
    counts = {}
    for job in jobs:
        key = job.get(field, "") or "Unknown"
        counts[key] = counts.get(key, 0) + 1
    for key, count in sorted(counts.items(), key=lambda x: x[1], reverse=True):
        ws.append([key, count])
    format_sheet(ws)


def write_excel(path, run_time, new_jobs, all_jobs, review, blocked, skipped, seen):
    wb = openpyxl.Workbook()
    write_jobs(wb.active, "New Jobs", new_jobs, seen, GREEN_FILL)

    combined = sorted(
        all_jobs + review + blocked,
        key=lambda job: (job["posted"], job["company"], job["title"]),
        reverse=True,
    )
    write_all_roles(wb.create_sheet(), combined, seen)

    write_jobs(wb.create_sheet(), "All Open Matches", all_jobs, seen)
    write_jobs(wb.create_sheet(), "Sponsorship Review", review, seen, YELLOW_FILL)
    write_jobs(wb.create_sheet(), "Explicitly Blocked", blocked, seen, RED_FILL)
    write_skipped(wb.create_sheet(), skipped)
    write_health(wb.create_sheet())
    write_counts(wb.create_sheet(), "By Company", all_jobs, "company")
    write_counts(wb.create_sheet(), "By ATS", all_jobs, "ats")
    write_counts(wb.create_sheet(), "By Location", all_jobs, "location")

    ws = wb.create_sheet("Run Summary")
    ws.append(["Metric", "Value"])
    ws.append(["Run Time UTC", run_time])
    ws.append(["New Jobs", len(new_jobs)])
    ws.append(["All Open Matches", len(all_jobs)])
    ws.append(["Sponsorship Review", len(review)])
    ws.append(["Explicitly Blocked", len(blocked)])
    ws.append(["Skipped Roles", len(skipped)])
    ws.append(["Scanner Errors", sum(row[2] == "error" for row in HEALTH)])
    format_sheet(ws)

    wb.save(path)


def write_markdown(path, run_time, new_jobs, review):
    lines = [
        "# Data Analytics / Data Engineering Job Digest",
        "",
        f"**Run:** {run_time}",
        f"**New matching jobs:** {len(new_jobs)}",
        f"**Needs sponsorship review:** {len(review)}",
        "",
        "## New Jobs",
        "",
    ]
    if not new_jobs:
        lines.append("No new matching roles found in this run.")
    else:
        for job in new_jobs:
            lines.extend([
                f"### [{job['title']}]({job['url']})",
                f"**{job['company']}** — {job['location'] or 'Location not listed'}  "
                f"| ATS: {job['ats']}  | Posted: {job['posted'] or 'Not listed'}  "
                f"| Sponsorship: {job['sponsorship_status']}",
                "",
            ])

    if review:
        lines.extend(["## Sponsorship / Export-Control Review", ""])
        for job in review:
            lines.append(
                f"- [{job['title']}]({job['url']}) — "
                f"{job['company']} — {job['location']} — "
                f"{job['sponsorship_reason']}"
            )

    path.write_text("\n".join(lines), encoding="utf-8")


def merge(source, target):
    for key in target:
        target[key].extend(source[key])


def main():
    RESULTS_DIR.mkdir(exist_ok=True)
    buckets = {"matched": [], "review": [], "blocked": [], "skipped": []}

    print(f"Scanning Greenhouse: {len(PRIORITY_GREENHOUSE)} companies")
    for slug in PRIORITY_GREENHOUSE:
        merge(scrape_greenhouse(slug), buckets)
        time.sleep(0.15)

    print(f"Scanning SmartRecruiters: {len(SMARTRECRUITERS_SLUGS)} companies")
    for slug in SMARTRECRUITERS_SLUGS:
        merge(scrape_smartrecruiters(slug), buckets)
        time.sleep(0.15)

    print(f"Scanning Workable: {len(WORKABLE_SLUGS)} companies")
    for slug in WORKABLE_SLUGS:
        merge(scrape_workable(slug), buckets)

    print(f"Scanning Workday Direct: {len(WORKDAY_DIRECT)} companies")
    for company in WORKDAY_DIRECT:
        merge(scrape_workday_direct(company), buckets)

    print(f"Scanning Workday Apify: {len(WORKDAY_APIFY)} companies")
    for company in WORKDAY_APIFY:
        merge(scrape_workday_apify(company), buckets)

    for key in buckets:
        buckets[key] = dedupe(buckets[key])

    all_jobs = sorted(
        buckets["matched"],
        key=lambda job: (job["posted"], job["company"], job["title"]),
        reverse=True,
    )

    run_time = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%MZ")
    seen = load_seen()
    new_jobs = []

    for job in all_jobs:
        job["is_new"] = job["id"] not in seen
        if job["is_new"]:
            seen[job["id"]] = run_time
            new_jobs.append(job)

    save_seen(seen)

    xlsx_path = RESULTS_DIR / f"data_jobs_{run_time}.xlsx"
    md_path = RESULTS_DIR / f"data_jobs_{run_time}.md"

    write_excel(
        xlsx_path, run_time, new_jobs, all_jobs,
        buckets["review"], buckets["blocked"], buckets["skipped"], seen,
    )
    write_markdown(md_path, run_time, new_jobs, buckets["review"])

    print(f"Done. New jobs: {len(new_jobs)}")
    print(f"All open matches: {len(all_jobs)}")
    print(f"Excel: {xlsx_path}")
    print(f"Markdown: {md_path}")


if __name__ == "__main__":
    main()
